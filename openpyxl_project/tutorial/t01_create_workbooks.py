from datetime import datetime
from openpyxl import Workbook  # type: ignore

wb = Workbook()

for sheet in wb:
    print(sheet.title)

ws = wb.active
ws["A1"] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted

ws["A3"] = datetime.now()

# Save the file
wb.save("sample.xlsx")
